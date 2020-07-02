import telebot
from telebot import apihelper
from telebot import types
from openpyxl import load_workbook
from telebot import apihelper
import re
import re
import os
from dotenv import load_dotenv


def parse_from_xlsx_data(data):
    pass


class QBlock:
    def __init__(self, bot_ref, args):
        title, questions_data = args
        self.bot = bot_ref
        self.questions = questions_data
        self.block_title = title
        self.points = 10
        self.answers = {j: 0 for j in tuple(map(lambda i: i[3], self.questions))}
        self.hypo_counter = 0

    def create_point_keyboard(self):
        markup = types.InlineKeyboardMarkup(row_width=4)
        tmp = [types.InlineKeyboardButton(text=str(i), callback_data=str(i)) for i in range(1, 5)]
        markup.add(*tmp)
        return markup

    def start(self, chat_id):
        text = '<b>{0}\n{1}</b> {2}'.format(self.block_title, self.questions[self.hypo_counter][0],
                                            self.questions[self.hypo_counter][2])
        self.bot.send_message(chat_id, text, parse_mode='html', reply_markup=self.create_point_keyboard())

    def write_answer(self, data):
        self.hypo_counter += 1
        if self.hypo_counter > 3:
            self.end_block()
        else:
            pass

    def send_question(self, chat_id):
        text = '<b>{0}\n{1}</b> {2}'.format(self.block_title, self.questions[self.hypo_counter][0],
                                            self.questions[self.hypo_counter][2])
        self.bot.send_message(chat_id, text, parse_mode='html', reply_markup=self.create_point_keyboard())

    def callback_handler(self, call):
        data = call.data.split('_')
        self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                   text=call.message.text, reply_markup=self.create_hypo_keyboard())
        self.bot.send_message(call.message.chat.id, 'Присвойте баллы данному утверждению от 2 до 10:',
                              reply_markup=self.create_point_keyboard())

    def end_block(self, chat_id):
        text = 'Блок {0} из 7 завершился. Идем дальше?'.format(self.block_title[5])
        markup = types.InlineKeyboardMarkup(1)
        markup.add(types.InlineKeyboardButton(text='Дальше', callback_data='NEXT_BLOCK'),
                   types.InlineKeyboardButton(text='Изменить', callback_data='CHANGE_BLOCK'))
        self.bot.send_message(chat_id, text, reply_markup=markup, parse_mode='html')

    def process_answer(self, call):
        markup = types.InlineKeyboardMarkup(4)
        buttons = ['1', '2', '3', '4']
        buttons[buttons.index(call.data)] = '✅ ' + buttons[buttons.index(call.data)]
        text = '<b>{0}</b> {1}'.format(self.questions[self.hypo_counter][0], self.questions[self.hypo_counter][2])
        markup.add(*[types.InlineKeyboardButton(text=i, callback_data='pa_{0}'.format(i)) for i in buttons])
        self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                   text=text, reply_markup=markup, parse_mode='html')
        data = int(call.data) - 1
        self.answers[self.questions[self.hypo_counter][3]] += data
        self.hypo_counter += 1
        if self.hypo_counter >= 6:
            self.end_block(call.message.chat.id)
        else:
            self.send_question(call.message.chat.id)

    def change(self, chat_id):
        self.answers = {j: 0 for j in tuple(map(lambda i: i[3], self.questions))}
        self.hypo_counter = 0
        self.start(chat_id)


class TeamworkTest:
    def __init__(self, filename, bot_ref):
        self.bot = bot_ref
        self.questions = self.create_questions_data(filename)
        self.blocks = self.create_blocks()
        self.block_counter = 0

    def handler(self, call):
        if len(re.findall(r'^0_[1-7]_[1-6]', call.data[:5])) == 1:
            if int(call.data[2]) - 1 == self.block_counter:
                self.return_cur_block().callback_handler(call)
            else:
                self.bot.send_message(call.message.chat.id, 'Пожалуйста, не надо так делать! Выбирайте варианты из '
                                                            'текущего блока :)')
        elif call.data in [str(i) for i in range(1, 5)]:
            try:
                self.return_cur_block().process_answer(call)
            except ValueError:
                self.bot.send_message(call.message.chat.id, 'Вы пытаетесь выйти за лимит 10 баллов на блок! Не надо '
                                                            'так :)')
        elif call.data == 'NEXT_BLOCK':
            self.block_counter += 1
            if self.block_counter <= len(self.blocks):
                try:
                    self.return_cur_block().start(call.message.chat.id)
                except AttributeError:
                    self.end_test(call.message.chat.id)
            else:
                self.end_test(call.message.chat.id)
        elif call.data == 'CHANGE_BLOCK':
            self.return_cur_block().change(call.message.chat.id)

    def create_blocks(self):
        res = []
        for i in range(1, 8):
            for j in self.questions.keys():
                if i == int(j):
                    res.append(QBlock(self.bot, *self.questions[j].items()))
                else:
                    continue
        return res

    @staticmethod
    def create_questions_data(filename):
        wb = load_workbook(filename=filename)
        sheet = wb.active
        alph = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        questions = dict.fromkeys(['1', '2', '3', '4', '5', '6', '7'], None)
        print(questions)
        tmp = None
        ranges = (sheet.max_row, sheet.max_column)
        for i in range(2, ranges[0] + 1):
            res = []
            for j in range(ranges[1] + 1):
                cur_place = '{0}{1}'.format(alph[j], i)
                try:
                    if len(sheet[cur_place].value) != 0:
                        res.append(sheet[cur_place].value)
                except TypeError:
                    continue
            if len(res) == 5:
                tmp = res[0]
                questions[tmp[5]] = {tmp: [res[1::]]}
            else:
                questions[tmp[5]][tmp].append(res)
        return questions

    def disclaimer(self, chat_id):
        text = 'Отлично!\n\n<b>Пожалуйста, внимательно прочитайте эту информацию:</b>\n\nТест состоит из 7 отдельных' \
               ' блоков по 6 утверждений.\n\nВам нужно выбрать степени согласия или несогласия с утверждениями, где:\n\n' \
               '4 - Полностью согласен\n3 - Скорее согласен\n2 - Скорее не согласен\n1 - Не согласен' \
               '\n\nВсе прочитали? ;)'
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(types.InlineKeyboardButton(text='Да, все понятно', callback_data='START_TEST'))
        self.bot.send_message(chat_id, text, reply_markup=markup, parse_mode='html')

    def start(self, chat_id):
        text = 'Приветствую!\n\n<b>Готовы пройти тест "Предпочитаемая роль в команде"?</b>\n\nЭто тест поможет ' \
               'выявить вашу предпочитаемую роль в команде.'
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(types.InlineKeyboardButton(text='Да, я хочу пройти', callback_data='CALL_DISC'))
        self.bot.send_message(chat_id, text, reply_markup=markup, parse_mode='html')

    def return_cur_block(self):
        try:
            return self.blocks[self.block_counter]
        except IndexError:
            pass

    def end_test(self, chat_id):
        final_res = self.calculate_result()
        answers = '\n'.join(['<b>{0} - {1}</b>'.format(''.join((i[0].upper(), i[1::])),
                                                       final_res[i]) for i in final_res.keys()])
        text = 'Спасибо за прохождение теста "Предпочитаемая роль в команде"!\n\n<b>Ваши результаты:\n\n{0}</b>\n\n' \
               'Успешной командной работы!'.format(answers)
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(types.InlineKeyboardButton(text='Хочу пройти еще раз', callback_data='RESTART_TEST'))
        self.bot.send_message(chat_id, text, reply_markup=markup, parse_mode='html')

    def calculate_result(self):
        res = dict.fromkeys(self.blocks[0].answers.keys(), 0)
        for i in self.blocks:
            for j in i.answers.keys():
                res[j] += i.answers[j]
        return res


class SimpleUserTestTeamwork:
    def __init__(self, user_id, bot_ref):
        self.on_test = False
        self.bot = bot_ref
        self.id = user_id
        self.test = TeamworkTest('Teamroles.xlsx', bot_ref)
        self.test.start(user_id)
        self.on_test = True

    def handler(self, message):
        if self.on_test:
            self.bot.send_message(message.chat.id, 'Вы уже проходите тест!')
        else:
            self.on_test = True
            self.test.start(message.chat.id)

    def callback_handler(self, call):
        if call.data == 'CALL_DISC':
            self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                       text=call.message.text, reply_markup=None, parse_mode='html')
            self.test.disclaimer(call.message.chat.id)

        elif call.data == 'START_TEST':
            self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                       text=call.message.text, reply_markup=None, parse_mode='html')
            self.test.return_cur_block().start(call.message.chat.id)

        elif call.data == 'RESTART_TEST':
            self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                       text=call.message.text, reply_markup=None, parse_mode='html')
            print(self.test)
            self.test = TeamworkTest('Teamroles.xlsx', self.bot)
            self.test.start(call.message.chat.id)
            print(self.test)
        else:
            self.test.handler(call)


if __name__ == '__main__':
    env_path = os.path.join(os.path.dirname(__file__), '.env')
    if os.path.exists(env_path):
        load_dotenv(env_path)
    API_TOKEN = os.getenv('API_TOKEN2')
    apihelper.proxy = {'https': 'http://admin@managewell.ru:gdkLH65%@ua23.nordvpn.com:80'}
    bot = telebot.TeleBot(API_TOKEN)
    users = {}

    @bot.message_handler(content_types='text')
    def handler(message):
        if message.text == '/start':
            try:
                users[message.chat.id].handler(message)
            except KeyError:
                users[message.chat.id] = SimpleUserTestTeamwork(message.chat.id, bot)
        else:
            bot.send_message(message.chat.id, 'Вам не нужно вводить никакой текст. Просто нажимайте кнопки! :)')

    @bot.callback_query_handler(func=lambda call: True)
    def smo(call):
        try:
            users[call.message.chat.id].callback_handler(call)
        except KeyError:
            users[call.message.chat.id] = SimpleUserTestTeamwork(call.message.chat.id, bot)

    bot.infinity_polling()
